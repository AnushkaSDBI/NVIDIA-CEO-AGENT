# ============================================================
#  src/to_excel.py  —  Make the corpus readable for review
#
#  Reads the cleaned documents (data/clean/documents.json) and
#  writes a formatted Excel workbook with two sheets:
#     1. Documents  — one row per document, filterable by source
#     2. Summary    — auto-counting per source + requirement check
#
#  Run:  python -m src.to_excel
# ============================================================

import os
import json
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import config as cfg

# openpyxl rejects ASCII control chars; PDF-extracted text often contains them.
try:
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE as _ILLEGAL
except Exception:
    _ILLEGAL = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")


def _xl(v):
    """Strip characters Excel can't store from any string cell value."""
    return _ILLEGAL.sub("", v) if isinstance(v, str) else v

OUT_PATH = "data/clean/documents.xlsx"

# Source categories we expect (used for the summary counts).
CATEGORIES = ["news", "company", "filing", "community", "research", "market", "pdf", "reference", "social"]

HEADER_FILL = PatternFill("solid", start_color="1F4E78")   # dark blue
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BASE_FONT   = Font(name="Arial")
TITLE_FONT  = Font(name="Arial", bold=True, size=14)


# --- 1. Load the cleaned documents --------------------------
def load_docs():
    with open(cfg.CLEAN_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


# --- 2. Sheet 1: Documents ----------------------------------
def build_documents_sheet(wb, docs):
    ws = wb.active
    ws.title = "Documents"

    headers = ["#", "Source", "Section", "Title", "Published", "URL", "Text (full)"]
    ws.append(headers)
    for cell in ws[1]:                              # style the header row
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    for i, d in enumerate(docs, start=1):
        ws.append([
            i,
            _xl(d.get("source", "")),
            _xl(d.get("section", "")),
            _xl(d.get("title", "")),
            _xl(d.get("published", "")),
            _xl(d.get("url", "")),
            _xl(d.get("text", "")[:32000]),         # full text (Excel cell limit ~32k)
        ])

    # Column widths for readability.
    widths = {"A": 6, "B": 13, "C": 24, "D": 55, "E": 22, "F": 50, "G": 80}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    for row in ws.iter_rows(min_row=2):             # base font on data rows
        for cell in row:
            cell.font = BASE_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"                          # keep header visible while scrolling
    ws.auto_filter.ref = f"A1:G{ws.max_row}"        # let the professor filter by source
    return ws


# --- 3. Sheet 2: Summary (live formulas) --------------------
def build_summary_sheet(wb):
    ws = wb.create_sheet("Summary")

    ws["A1"] = f"AI CEO — Collected Documents Summary ({cfg.COMPANY['name']})"
    ws["A1"].font = TITLE_FONT

    ws["A3"], ws["B3"] = "Company:", cfg.COMPANY["name"]
    ws["A4"], ws["B4"] = "Generated:", datetime.now().strftime("%Y-%m-%d %H:%M")

    ws["A6"], ws["B6"] = "Source", "Count"
    for cell in (ws["A6"], ws["B6"]):
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT

    # COUNTIF formulas reference the Documents sheet -> stay live.
    row = 7
    for cat in CATEGORIES:
        ws[f"A{row}"] = cat
        ws[f"B{row}"] = f'=COUNTIF(Documents!B:B,"{cat}")'
        row += 1

    ws[f"A{row}"], ws[f"B{row}"] = "Total documents", "=COUNT(Documents!A:A)"
    total_row = row
    row += 1
    ws[f"A{row}"], ws[f"B{row}"] = "Sources used", f'=COUNTIF(B7:B{total_row-1},">0")'
    used_row = row
    row += 1
    ws[f"A{row}"] = "Requirement (>=100 docs & >=3 sources)"
    ws[f"B{row}"] = f'=IF(AND(B{total_row}>=100,B{used_row}>=3),"PASS","CHECK")'

    for r in ws.iter_rows():                        # apply base font everywhere
        for cell in r:
            if cell.font is BASE_FONT or cell.value is not None and cell.row > 1:
                cell.font = BASE_FONT
    ws["A1"].font = TITLE_FONT                       # restore title style

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    return ws


# --- Sheet 3: Stock (prices + technical indicators) ---------
def build_stock_sheet(wb):
    import os
    path = "data/clean/stock.json"
    if not os.path.exists(path):
        return
    with open(path, encoding="utf-8") as f:
        stock = json.load(f)
    ws = wb.create_sheet("Stock")

    ws["A1"] = f"{cfg.COMPANY['name']} ({stock.get('ticker','')}) — Stock & Technical Analysis"
    ws["A1"].font = TITLE_FONT

    # Indicators block
    ws["A3"] = "Indicator"; ws["B3"] = "Value"
    for c in (ws["A3"], ws["B3"]):
        c.fill, c.font = HEADER_FILL, HEADER_FONT
    r = 4
    for k, v in stock.get("indicators", {}).items():
        ws[f"A{r}"] = k; ws[f"B{r}"] = v; r += 1

    # Price history header (a few rows below)
    start = r + 2
    ws[f"A{start}"] = "Date"; ws[f"B{start}"] = "Close"; ws[f"C{start}"] = "Volume"
    for col in ("A", "B", "C"):
        ws[f"{col}{start}"].fill = HEADER_FILL
        ws[f"{col}{start}"].font = HEADER_FONT
    for i, p in enumerate(stock.get("history", []), start=start + 1):
        ws[f"A{i}"] = p["date"]; ws[f"B{i}"] = p["close"]; ws[f"C{i}"] = p["volume"]

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    return ws


# --- 4. Orchestrate -----------------------------------------
def run():
    docs = load_docs()
    wb = Workbook()
    build_documents_sheet(wb, docs)
    build_summary_sheet(wb)
    build_stock_sheet(wb)
    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {len(docs)} documents to {OUT_PATH}")


if __name__ == "__main__":
    run()
